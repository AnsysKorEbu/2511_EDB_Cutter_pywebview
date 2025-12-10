# preprocessing.py
"""
Data preprocessing functions for stackup processing.
Handles Excel data extraction, validation, and transformation.
"""
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from util.logger_module import logger
import os
import sys
from stackup.utils.excel_utils import ExcelUtilities

# Optional PySide6 import for GUI dialogs
try:
    from PySide6.QtWidgets import QApplication, QMessageBox
    HAS_PYSIDE6 = True
except ImportError:
    HAS_PYSIDE6 = False


def find_real_value(ws, row, col):
    for merged_range in ws.merged_cells.ranges:
        if (row, col) in merged_range.cells:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return ws.cell(row=row, column=col).value


def sanitize_name(net):
    if net is None:
        return net
    net = net.replace('-', 'm')
    net = net.replace('.', 'p')
    net = net.replace(' ', '_')
    net = net.replace('/', '_')
    return net


def get_net_color(net_name):
    """Get color for a net based on its name, similar to q2d_worker's logic."""
    net_name = net_name.upper()
    if any(x in net_name for x in ['VBATT']):
        return 	(255, 60, 60)
    if 'VBUS' in net_name:
        return (255, 192, 203)
    if 'DISPLAY_GND' in net_name:
        return (135, 206, 235)
    if 'EMI' in net_name.upper():
        return 	(0, 0, 0)
    if any(x in net_name for x in ['WPC', 'NFC', 'SPK', 'MOT']):
        return 	(168, 60, 168)
    if 'FRC' in net_name or any(x in net_name.split('_') for x in ['LB', 'MB', 'HB', 'UHB']):
        return 	(152, 255, 152)
    if 'LVDS' in net_name or net_name.split('_')[-1] in ['P', 'N', 'M']:
        return (0, 128, 0)
    if any(x in net_name for x in ['DISPLAY_VDD', 'VSS', 'VDD']):
        return 	(255, 185, 40)
    if net_name in ['SPACE', 'S']:
        return (221, 221, 221)
    if any(x in net_name for x in ['GND', 'GROUND']):
        return 	(80, 80, 255)
    return (190, 190, 190) # Default blue for normal nets


def check_file_exists(filename):
    """
    Check if a file exists and show a popup warning if it doesn't.

    Args:
        filename (str): Path to the file to check

    Returns:
        bool: True if file exists, False otherwise
    """
    if os.path.isfile(filename):
        return True

    # Log the error
    file_name = os.path.basename(filename)
    logger.error(f"Required file not found: {filename}")

    # Show GUI dialog if PySide6 is available
    if HAS_PYSIDE6:
        try:
            # Make sure we have a QApplication instance
            app = QApplication.instance()
            if not app:
                app = QApplication(sys.argv)

            # Show error message
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle("Alert")
            msg.setText(f"필요한 파일을 찾을 수 없습니다: {file_name}")
            msg.setInformativeText(f"자세한 사항은 config.py와 config.json을 확인하세요.")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.button(QMessageBox.Ok).setText("종료")
            msg.exec()
        except Exception as e:
            logger.warning(f"Could not show GUI dialog: {e}")

    return False


def read_net_width(filename, sheet_name, start_cell, material):
    if not check_file_exists(filename):
        sys.exit(1)

    wb = load_workbook(filename, data_only=True)
    ws = wb[sheet_name]
    row, col = coordinate_to_tuple(start_cell)
    result = []
    while True:
        net = ws.cell(row=row, column=col).value
        if net is None:
            break
        if net == 'total width':
            width = ws.cell(row=row + 1, column=col).value
            result.append({'net': net, 'width': width, 'material': 'air', 'color': (128, 128, 128)})
            break
        if net in ('s', 'space') or net is None:
            net = 'space'
        net = sanitize_name(net)
        width = ws.cell(row=row + 1, column=col).value
        if width is None:
            col += 1
            continue

        # Get color for this net
        color = get_net_color(net)

        result.append({'net': net, 'width': width, 'material': material, 'color': color})
        col += 1
    return result


def read_material_properties(filename):
    if not check_file_exists(filename):
        sys.exit(1)

    wb = load_workbook(filename, data_only=True)
    ws = wb.worksheets[0]
    current_row = 3
    result = []

    while True:
        check_value = ws.cell(row=current_row, column=4).value
        if check_value is None:
            break

        key_name = find_real_value(ws, current_row, 2)
        material_name = find_real_value(ws, current_row, 4)
        cu_foil = find_real_value(ws, current_row, 8)
        dk_df = find_real_value(ws, current_row, 9)
        height = find_real_value(ws, current_row, 30)

        if height is None:
            current_row += 1
            continue

        if cu_foil == '-':
            cu_foil = None

        entry = {
            'layer': key_name,
            'row': current_row,
            'material': sanitize_name(material_name),
            'CU_foil': cu_foil,
            'Dk/Df': dk_df,
            'height': height
        }

        result.append(entry)

        if material_name and material_name.strip().lower() == 'adhesive':
            air_gap_height = calculate_adhesive_air_gap(ws, current_row, height)

            if air_gap_height is not None:
                air_entry = {
                    'layer': key_name,
                    'row': current_row,
                    'material': 'air',
                    'CU_foil': None,
                    'Dk/Df': '1.006/0(10GHz)',
                    'height': air_gap_height
                }
                result.append(air_entry)

        current_row += 1

    return result


def calculate_adhesive_air_gap(ws, current_row, height):
    """
    Calculate air gap for Adhesive materials by finding max height from other layers.

    Args:
        ws: worksheet object
        current_row: current row number
        height: current layer height

    Returns:
        float or None: air gap value if calculated, None otherwise
    """
    other_heights = []

    for col in range(9, 30):
        if col == 29:
            continue

        cell_value = find_real_value(ws, current_row, col)

        if cell_value is not None:
            try:
                numeric_value = float(cell_value)
                other_heights.append(numeric_value)
            except (ValueError, TypeError):
                continue

    if other_heights:
        max_other_height = max(other_heights)
        if max_other_height > height:
            return max_other_height - height

    return None


def read_layer_material(filename):
    if not check_file_exists(filename):
        sys.exit(1)

    wb = load_workbook(filename, data_only=True)
    ws = wb.worksheets[0]
    current_row = 3
    result = []
    while True:
        if ws.cell(row=current_row, column=4).value is None:
            break
        key_name = find_real_value(ws, current_row, 2)
        material_name = find_real_value(ws, current_row, 4)
        entry = {
            'layer': key_name,
            'row': current_row,
            'material': material_name
        }
        result.append(entry)
        current_row += 1
    return result


def find_center_layer(entries):
    # Extract layer numbers, converting strings like '1LAYER' to integers
    layers = []
    for e in entries:
        if e['layer'] is not None:
            layer_val = e['layer']
            # Convert string layer names to integers
            if isinstance(layer_val, str):
                try:
                    # Extract number from strings like '1LAYER'
                    layer_num = int(layer_val.upper().replace('LAYER', '').strip())
                    layers.append(layer_num)
                except ValueError:
                    continue
            elif isinstance(layer_val, (int, float)):
                layers.append(int(layer_val))

    if not layers:
        return None

    mid = max(layers) // 2
    for i, entry in enumerate(entries):
        if i == 0 or i >= len(entries) - 1:
            continue

        prev_layer = entries[i - 1]['layer']
        next_layer = entries[i + 1]['layer']

        # Convert to int for comparison
        try:
            if isinstance(prev_layer, str):
                prev_layer = int(prev_layer.upper().replace('LAYER', '').strip())
            if isinstance(next_layer, str):
                next_layer = int(next_layer.upper().replace('LAYER', '').strip())

            if prev_layer == mid and next_layer == mid + 1:
                return entry['row']
        except (ValueError, AttributeError):
            continue

    return None


def get_max_total_width(net_data):
    widths = [
        entry['width']
        for entries in net_data.values()
        for entry in entries
        if entry.get('net') == 'total width'
    ]
    return max(widths) if widths else None


def swap_all_air_adhesive_pairs(layer_data, center_row):
    """
    Swaps all air materials with their preceding Adhesive materials
    if air's row is smaller than center_row
    """

    swapped_count = 0

    for i in range(1, len(layer_data)):
        current_item = layer_data[i]
        previous_item = layer_data[i - 1]

        if (current_item.get('material') == 'air' and
                current_item.get('row', float('inf')) < center_row and
                'adhesive' in previous_item.get('material').lower()):
            layer_data[i], layer_data[i - 1] = layer_data[i - 1], layer_data[i]

            logger.info(f"Swapped air (row {current_item.get('row')}) with Adhesive")
            swapped_count += 1

    logger.info(f"Total swaps performed: {swapped_count}")
    return layer_data


def check_sheet_exists(excel_file, sheet_name):
    """
    Check if sheet exists in excel file
    Uses ExcelUtilities for consistent implementation
    """
    return ExcelUtilities.check_sheet_exists(excel_file, sheet_name)


def find_text_cell_in_column_c(excel_file, sheet_name):
    """
    Find first text cell in column C of given sheet
    Returns cell address like 'C4', 'C2', etc.
    Uses ExcelUtilities for consistent implementation
    """
    return ExcelUtilities.find_text_cell_in_column(excel_file, sheet_name, 'C')


def process_sheets_with_validation_backup(excel_file, sheet_names):
    """
    Process all sheets with existence check and cell finding
    """
    net_data_origin = {}

    for sheet in sheet_names:
        if not check_sheet_exists(excel_file, sheet):
            net_data_origin[sheet] = []
            continue

        cell = find_text_cell_in_column_c(excel_file, sheet)
        if cell:
            net_data_origin[sheet] = read_net_width(excel_file, sheet, cell, 'copper')
        else:
            net_data_origin[sheet] = []

    return net_data_origin

def process_sheets_with_validation(excel_file, sheet_names):
    """
    Process all sheets with existence check and cell finding
    Filters out 'total width' entries from results
    """
    net_data_origin = {}

    for sheet in sheet_names:
        if not check_sheet_exists(excel_file, sheet):
            net_data_origin[sheet] = []
            continue

        cell = find_text_cell_in_column_c(excel_file, sheet)
        if cell:
            raw_data = read_net_width(excel_file, sheet, cell, 'copper')
            filtered_data = [entry for entry in raw_data if entry['net'] != 'total width']
            net_data_origin[sheet] = filtered_data
        else:
            net_data_origin[sheet] = []

    return net_data_origin
