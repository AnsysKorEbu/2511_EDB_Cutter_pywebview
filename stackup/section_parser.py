"""
Section Parser Module

Extracts section names from Excel stackup files based on LAYER pattern.
Scans Excel file to find cells containing "LAYER" and extracts section names
from adjacent columns following the pattern:
  - Column N: "LAYER" (anchor)
  - Column N+1: spec value
  - Column N+2: thick value
  - Column N+3: +- tolerance value
  - Column N+4: section name (e.g., "C/N 1", "C/N 1-1")
"""

import os
from pathlib import Path
from typing import List, Tuple, Optional, Dict
from openpyxl import load_workbook
from util.logger_module import logger


def extract_section_names(excel_path: str) -> Dict:
    """
    Extract section names and column numbers from Excel stackup file using LAYER pattern.

    Scans the Excel file from row 1 downward, looking for cells containing "LAYER".
    When found, extracts the section name from 4 columns to the right along with column number.

    Args:
        excel_path: Absolute path to Excel file (.xlsx or .xls)

    Returns:
        Dictionary with structure:
        {
            'success': bool,
            'sections': List[str],  # List of section names (e.g., ["C/N 1", "C/N 1-1"])
            'section_columns': Dict[str, int],  # Mapping of section name to column number
            'file_path': str,       # Absolute path to Excel file
            'error': str            # Error message if failed (only if success=False)
        }

    Example:
        >>> result = extract_section_names("C:/path/to/stackup.xlsx")
        >>> if result['success']:
        >>>     print(result['sections'])  # ["C/N 1", "C/N 1-1", "C/N 2"]
        >>>     print(result['section_columns'])  # {"C/N 1": 94, "C/N 1-1": 110, ...}
    """
    try:
        # Validate file exists
        if not os.path.exists(excel_path):
            error_msg = f"Excel file not found: {excel_path}"
            logger.error(error_msg)
            return {
                'success': False,
                'sections': [],
                'file_path': excel_path,
                'error': error_msg
            }

        # Validate file extension
        file_ext = Path(excel_path).suffix.lower()
        if file_ext not in ['.xlsx', '.xls']:
            error_msg = f"Invalid file type: {file_ext}. Expected .xlsx or .xls"
            logger.error(error_msg)
            return {
                'success': False,
                'sections': [],
                'file_path': excel_path,
                'error': error_msg
            }

        logger.info(f"Loading Excel file: {excel_path}")

        # Load workbook
        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active  # Use active sheet
        except Exception as e:
            error_msg = f"Failed to load Excel file: {str(e)}"
            logger.error(error_msg)
            return {
                'success': False,
                'sections': [],
                'file_path': excel_path,
                'error': error_msg
            }

        logger.info(f"Scanning for LAYER pattern (sheet: {ws.title}, rows: {ws.max_row})")

        # Find all LAYER cells
        layer_cells = _find_layer_cells(ws)

        if not layer_cells:
            warning_msg = "No cells containing 'LAYER' found in Excel file"
            logger.warning(warning_msg)
            return {
                'success': True,  # Not an error, just no sections found
                'sections': [],
                'file_path': excel_path,
                'error': warning_msg
            }

        logger.info(f"Found {len(layer_cells)} LAYER cells")

        # Extract section names and column numbers from LAYER cells
        sections = []
        section_columns = {}  # Map section name to column number
        for row, col in layer_cells:
            row_sections, row_columns = _extract_section_from_cell(ws, row, col)
            sections.extend(row_sections)  # Add all sections from this row
            section_columns.update(row_columns)  # Add column mappings

        # Filter out explicitly excluded section names only
        filtered_sections = []
        filtered_section_columns = {}

        for section in sections:
            # Skip explicitly excluded names (like NAMES OF GOODS)
            if section in {'NAMES OF GOODS', 'NAME OF GOODS'}:
                logger.debug(f"Filtered out excluded section: '{section}'")
                continue

            # Keep all other sections
            filtered_sections.append(section)
            # Keep column mapping for this section
            if section in section_columns:
                filtered_section_columns[section] = section_columns[section]

        # Deduplicate while preserving order
        unique_sections = []
        unique_section_columns = {}
        seen = set()
        for section in filtered_sections:
            if section not in seen:
                unique_sections.append(section)
                seen.add(section)
                # Keep first occurrence's column number
                if section in filtered_section_columns:
                    unique_section_columns[section] = filtered_section_columns[section]

        logger.info(f"Extracted {len(unique_sections)} unique sections: {unique_sections}")
        logger.info(f"Section column mappings: {unique_section_columns}")

        wb.close()

        return {
            'success': True,
            'sections': unique_sections,
            'section_columns': unique_section_columns,
            'file_path': excel_path
        }

    except Exception as e:
        error_msg = f"Unexpected error extracting sections: {str(e)}"
        logger.error(error_msg)
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'sections': [],
            'section_columns': {},
            'file_path': excel_path,
            'error': error_msg
        }


def _find_layer_cells(ws) -> List[Tuple[int, int]]:
    """
    Find the FIRST cell containing exactly "LAYER" text (case-insensitive).

    Args:
        ws: openpyxl worksheet object

    Returns:
        List with single (row, col) tuple for first "LAYER" cell found

    Example:
        >>> layer_cells = _find_layer_cells(ws)
        >>> # [(5, 2)]  # First LAYER found in column 2 at row 5
    """
    # Scan all rows
    for row in range(1, ws.max_row + 1):
        # Scan all columns (up to reasonable limit to avoid excessive scanning)
        for col in range(1, min(ws.max_column + 1, 100)):  # Limit to first 100 columns
            cell_value = ws.cell(row=row, column=col).value

            # Check if cell is exactly "LAYER" (case-insensitive, exact match only)
            if cell_value and isinstance(cell_value, str):
                if cell_value.strip().upper() == 'LAYER':
                    logger.debug(f"Found first LAYER at ({row}, {col}): '{cell_value}'")
                    return [(row, col)]  # Return immediately after finding first LAYER

    return []  # No LAYER found


def _extract_section_from_cell(ws, row: int, col: int) -> Tuple[List[str], Dict[str, int]]:
    """
    Extract ALL non-empty values from the row after LAYER cell with their column numbers.

    Scans the entire row from LAYER column onwards and collects all
    non-empty cell values as potential section names along with their column numbers.

    Args:
        ws: openpyxl worksheet
        row: Row number (1-indexed) where "LAYER" was found
        col: Column number (1-indexed) where "LAYER" was found

    Returns:
        Tuple of (section_names, section_columns):
            - section_names: List of potential section names
            - section_columns: Dict mapping section name to column number

    Example:
        >>> sections, columns = _extract_section_from_cell(ws, 8, 2)
        >>> # sections = ["SPEC_1", "12.5um", "C/N 1", "C/N 1-1", ...]
        >>> # columns = {"C/N 1": 94, "C/N 1-1": 110, ...}
    """
    sections = []
    section_columns = {}

    try:
        # Scan from LAYER column + 1 to end of row
        max_col = ws.max_column  # Check all columns to the end

        for scan_col in range(col + 1, max_col + 1):
            cell_value = ws.cell(row=row, column=scan_col).value

            # Skip empty cells
            if not cell_value:
                continue

            # Skip the LAYER text itself (in case it appears again)
            if isinstance(cell_value, str) and 'LAYER' in cell_value.upper():
                continue

            # Convert to string, replace newlines with spaces, and trim
            value_str = str(cell_value).replace('\n', ' ').replace('\r', ' ').strip()

            if value_str:
                sections.append(value_str)
                section_columns[value_str] = scan_col  # Store column number
                logger.debug(f"Found potential section at ({row}, {scan_col}): '{value_str}'")

        return sections, section_columns

    except Exception as e:
        logger.debug(f"Error extracting sections from ({row}, {col}): {e}")
        return [], {}


def validate_section_name(section_name) -> bool:
    """
    Validate that a section name is valid (not None, not empty, is string).

    Args:
        section_name: Section name to validate

    Returns:
        bool: True if valid section name, False otherwise

    Example:
        >>> validate_section_name("C/N 1")     # True
        >>> validate_section_name("")          # False
        >>> validate_section_name(None)        # False
        >>> validate_section_name(123)         # False (not string)
    """
    if section_name is None:
        return False

    if not isinstance(section_name, str):
        return False

    # Remove whitespace and check if empty
    if not section_name.strip():
        return False

    return True


if __name__ == "__main__":
    # Test the section parser with sample file
    import sys

    if len(sys.argv) > 1:
        test_file = sys.argv[1]
    else:
        # Use default test file
        test_file = Path(__file__).parent / "rawdata.xlsx"

    print(f"Testing section parser with: {test_file}")
    result = extract_section_names(str(test_file))

    print("\n" + "="*60)
    print("RESULT:")
    print("="*60)
    print(f"Success: {result['success']}")
    print(f"File: {result['file_path']}")

    if result['success']:
        print(f"Sections found ({len(result['sections'])}):")
        for i, section in enumerate(result['sections'], 1):
            print(f"  {i}. {section}")
    else:
        print(f"Error: {result.get('error', 'Unknown error')}")
