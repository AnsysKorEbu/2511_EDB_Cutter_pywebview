"""
Stackup Data Export Module

Exports stackup data from rawdata.xlsx to user-friendly Excel format.
Provides modular functions for data extraction, formatting, and file generation.
"""

import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from util.logger_module import logger


# Constants
HEADER_FILL_COLOR = "667eea"  # Primary purple from GUI theme
ALT_ROW_FILL_COLOR = "2d2d30"  # Alternating row color (dark theme)
HEADER_FONT_COLOR = "ffffff"  # White text for headers
EXPORTER_VERSION = "1.0"


def export_stackup_to_excel(source_excel_path: str, output_excel_path: str, config: Optional[Dict] = None) -> Dict:
    """
    Extract stackup data from source Excel and export to formatted Excel file.

    Main orchestration function that coordinates data extraction and Excel generation.

    Args:
        source_excel_path: Path to rawdata.xlsx file
        output_excel_path: Path for output Excel file
        config: Optional configuration dict containing:
                - cut_stackup_mapping: Dict mapping cut IDs to section names
                - section_columns: Dict mapping section names to column numbers

    Returns:
        Dictionary with keys:
            - success (bool): Whether export succeeded
            - output_path (str): Path to generated file (if successful)
            - rows_exported (int): Number of data rows exported (if successful)
            - error (str): Error message (if failed)
    """
    try:
        logger.info(f"Starting stackup export from: {source_excel_path}")
        logger.info(f"Output file: {output_excel_path}")

        # Validate source file exists
        if not os.path.exists(source_excel_path):
            error_msg = f"Source Excel file not found: {source_excel_path}"
            logger.error(error_msg)
            return {'success': False, 'error': error_msg}

        # Extract stackup data using existing module
        # Note: Wrapped in try-except because read_material_properties calls sys.exit on error
        try:
            # Check if config with cut/section mapping is provided
            if config and config.get('cut_stackup_mapping') and config.get('section_columns'):
                logger.info("Extracting data for each cut/section pair")
                stackup_data = _extract_stackup_data_by_sections(source_excel_path, config)
            else:
                logger.info("Extracting data using default column")
                stackup_data = _extract_stackup_data(source_excel_path)

            raw_materials = _extract_raw_materials(source_excel_path)
        except SystemExit as e:
            error_msg = f"Failed to extract stackup data: System exit called (code: {e.code})"
            logger.error(error_msg)
            return {'success': False, 'error': error_msg}
        except Exception as e:
            error_msg = f"Failed to extract stackup data: {str(e)}"
            logger.error(error_msg)
            return {'success': False, 'error': error_msg}

        if not stackup_data:
            error_msg = "No stackup data extracted from source file"
            logger.warning(error_msg)
            return {'success': False, 'error': error_msg}

        # Create Excel workbook with 3 sheets
        try:
            wb = Workbook()

            # Sheet 1: Stackup Summary (processed data)
            ws_summary = wb.active
            ws_summary.title = "Stackup Summary"
            format_stackup_summary_sheet(ws_summary, stackup_data)

            # Sheet 2: Raw Materials (original specifications)
            ws_raw = wb.create_sheet("Raw Materials")
            format_raw_materials_sheet(ws_raw, raw_materials)

            # Sheet 3: Export Info (metadata)
            ws_info = wb.create_sheet("Export Info")
            export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            add_export_metadata_sheet(ws_info, source_excel_path, export_time, len(stackup_data))

            # Save workbook
            wb.save(output_excel_path)
            logger.info(f"Successfully exported {len(stackup_data)} rows to: {output_excel_path}")

            return {
                'success': True,
                'output_path': output_excel_path,
                'rows_exported': len(stackup_data)
            }

        except PermissionError as e:
            error_msg = f"Permission denied writing to file: {output_excel_path}"
            logger.error(f"{error_msg} - {str(e)}")
            return {'success': False, 'error': error_msg}
        except Exception as e:
            error_msg = f"Failed to write output file: {str(e)}"
            logger.error(error_msg)
            return {'success': False, 'error': error_msg}

    except Exception as e:
        error_msg = f"Unexpected error during export: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return {'success': False, 'error': error_msg}


def _extract_stackup_data(source_excel_path: str) -> List[Dict]:
    """
    Extract processed stackup data from source Excel file.

    Calls excel_reader.read_material_properties() to get processed data.
    Wrapped to handle potential sys.exit() calls.

    Args:
        source_excel_path: Path to rawdata.xlsx

    Returns:
        List of dictionaries with stackup properties:
            - layer: Layer identifier (e.g., "1LAYER", "1.5LAYER", "3L-1LAYER")
            - row: Row number
            - material: Material name (e.g., "copper", "film", "adhesive")
            - CU_foil: Copper foil type (if applicable)
            - Dk/Df: Dielectric constant/dissipation factor
            - height: Thickness in micrometers
    """
    from stackup.excel_reader import read_material_properties

    logger.info("Extracting stackup data using excel_reader.read_material_properties()")
    return read_material_properties(source_excel_path)


def _extract_stackup_data_by_sections(source_excel_path: str, config: Dict) -> List[Dict]:
    """
    Extract stackup data for each cut/section pair using section-specific columns.

    Iterates through cut_stackup_mapping and extracts data using the corresponding
    section's column number from section_columns.

    Args:
        source_excel_path: Path to rawdata.xlsx
        config: Configuration dict containing:
                - cut_stackup_mapping: Dict mapping cut IDs to section names
                - section_columns: Dict mapping section names to column numbers

    Returns:
        List of dictionaries with stackup properties including:
            - cut_id: Cut identifier (e.g., "cut_001")
            - section: Section name (e.g., "C/N 1")
            - layer: Layer identifier
            - row: Row number
            - material: Material name
            - CU_foil: Copper foil type (if applicable)
            - Dk/Df: Dielectric constant/dissipation factor
            - height: Thickness in micrometers
    """
    from stackup.excel_reader import read_material_properties

    cut_stackup_mapping = config.get('cut_stackup_mapping', {})
    section_columns = config.get('section_columns', {})

    logger.info(f"Extracting data for {len(cut_stackup_mapping)} cut/section pairs")

    all_data = []

    for cut_id, section_name in cut_stackup_mapping.items():
        # Get column number for this section
        section_column = section_columns.get(section_name)

        if section_column is None:
            logger.warning(f"No column found for section '{section_name}', skipping cut '{cut_id}'")
            continue

        logger.info(f"Extracting data for {cut_id} -> {section_name} (column {section_column})")

        # Extract data using section-specific column
        try:
            section_data = read_material_properties(source_excel_path, section_column)

            # Add cut_id and section to each row
            for row in section_data:
                row['cut_id'] = cut_id
                row['section'] = section_name

            all_data.extend(section_data)
            logger.info(f"  Extracted {len(section_data)} rows for {cut_id}")

        except Exception as e:
            logger.error(f"Failed to extract data for {cut_id} ({section_name}): {str(e)}")
            continue

    logger.info(f"Total rows extracted: {len(all_data)}")
    return all_data


def _extract_raw_materials(source_excel_path: str) -> Dict:
    """
    Extract raw material specifications from source Excel file.

    Calls excel_reader.extract_materials_specifications() to get original specs.

    Args:
        source_excel_path: Path to rawdata.xlsx

    Returns:
        Dictionary with keys:
            - main_materials: Dict of main material specifications
            - subsidiary_materials: Dict of subsidiary material specifications
    """
    from openpyxl import load_workbook
    from stackup.excel_reader import extract_materials_specifications

    logger.info("Extracting raw materials using excel_reader.extract_materials_specifications()")

    try:
        wb = load_workbook(source_excel_path, data_only=True)
        ws = wb.worksheets[0]
        return extract_materials_specifications(ws)
    except Exception as e:
        logger.warning(f"Failed to extract raw materials: {str(e)}")
        return {'main_materials': {}, 'subsidiary_materials': {}}


def format_stackup_summary_sheet(worksheet: Worksheet, stackup_data: List[Dict]) -> None:
    """
    Format the Stackup Summary sheet with processed stackup data.

    Creates a table with headers and styled data rows.
    Applies alternating row colors and auto-adjusts column widths.

    Args:
        worksheet: openpyxl Worksheet object to format
        stackup_data: List of stackup property dictionaries
    """
    logger.info(f"Formatting Stackup Summary sheet with {len(stackup_data)} rows")

    # Check if data includes cut_id and section (multi-section export)
    has_cut_section = False
    if stackup_data and 'cut_id' in stackup_data[0] and 'section' in stackup_data[0]:
        has_cut_section = True
        logger.info("Data includes Cut ID and Section columns")

    # Define headers based on data type
    if has_cut_section:
        headers = ["Cut ID", "Section", "Layer", "Row", "Material", "CU Foil", "Dk/Df", "Height (um)"]
    else:
        headers = ["Layer", "Row", "Material", "CU Foil", "Dk/Df", "Height (um)"]

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        _apply_header_style(cell)

    # Write data rows
    for row_idx, data_row in enumerate(stackup_data, start=2):
        if has_cut_section:
            # Multi-section export with Cut ID and Section
            worksheet.cell(row=row_idx, column=1, value=data_row.get('cut_id'))
            worksheet.cell(row=row_idx, column=2, value=data_row.get('section'))
            worksheet.cell(row=row_idx, column=3, value=data_row.get('layer'))
            worksheet.cell(row=row_idx, column=4, value=data_row.get('row'))
            worksheet.cell(row=row_idx, column=5, value=data_row.get('material'))
            worksheet.cell(row=row_idx, column=6, value=data_row.get('CU_foil'))
            worksheet.cell(row=row_idx, column=7, value=data_row.get('Dk/Df'))
            worksheet.cell(row=row_idx, column=8, value=data_row.get('height'))
        else:
            # Standard export
            worksheet.cell(row=row_idx, column=1, value=data_row.get('layer'))
            worksheet.cell(row=row_idx, column=2, value=data_row.get('row'))
            worksheet.cell(row=row_idx, column=3, value=data_row.get('material'))
            worksheet.cell(row=row_idx, column=4, value=data_row.get('CU_foil'))
            worksheet.cell(row=row_idx, column=5, value=data_row.get('Dk/Df'))
            worksheet.cell(row=row_idx, column=6, value=data_row.get('height'))

        # Apply alternating row color (every other row)
        if row_idx % 2 == 0:
            _apply_alt_row_style(worksheet, row_idx, len(headers))

    # Auto-adjust column widths
    _auto_adjust_column_widths(worksheet, headers)

    logger.info("Stackup Summary sheet formatted successfully")


def format_raw_materials_sheet(worksheet: Worksheet, materials_dict: Dict) -> None:
    """
    Format the Raw Materials sheet with original material specifications.

    Creates two sections:
    1. Main Materials (FCCL, COVERLAY, PREPREG)
    2. Subsidiary Materials (PSR, STIFFENER, INK)

    Args:
        worksheet: openpyxl Worksheet object to format
        materials_dict: Dictionary with 'main_materials' and 'subsidiary_materials' keys
    """
    logger.info("Formatting Raw Materials sheet")

    current_row = 1

    # Section 1: Main Materials
    main_materials = materials_dict.get('main_materials', {})
    if main_materials:
        # Section header
        cell = worksheet.cell(row=current_row, column=1, value="Main Materials")
        cell.font = Font(bold=True, size=14)
        current_row += 1

        # Column headers
        worksheet.cell(row=current_row, column=1, value="Material Type")
        worksheet.cell(row=current_row, column=2, value="Specification")
        _apply_header_style(worksheet.cell(row=current_row, column=1))
        _apply_header_style(worksheet.cell(row=current_row, column=2))
        current_row += 1

        # Data rows
        for material_type, specification in main_materials.items():
            worksheet.cell(row=current_row, column=1, value=material_type)
            worksheet.cell(row=current_row, column=2, value=specification)
            current_row += 1

        current_row += 1  # Blank row separator

    # Section 2: Subsidiary Materials
    subsidiary_materials = materials_dict.get('subsidiary_materials', {})
    if subsidiary_materials:
        # Section header
        cell = worksheet.cell(row=current_row, column=1, value="Subsidiary Materials")
        cell.font = Font(bold=True, size=14)
        current_row += 1

        # Column headers
        worksheet.cell(row=current_row, column=1, value="Material Type")
        worksheet.cell(row=current_row, column=2, value="Specification")
        _apply_header_style(worksheet.cell(row=current_row, column=1))
        _apply_header_style(worksheet.cell(row=current_row, column=2))
        current_row += 1

        # Data rows
        for material_type, specification in subsidiary_materials.items():
            worksheet.cell(row=current_row, column=1, value=material_type)
            worksheet.cell(row=current_row, column=2, value=specification)
            current_row += 1

    # Auto-adjust column widths
    _auto_adjust_column_widths(worksheet, ["Material Type", "Specification"])

    logger.info("Raw Materials sheet formatted successfully")


def add_export_metadata_sheet(
    worksheet: Worksheet,
    source_path: str,
    export_time: str,
    row_count: int
) -> None:
    """
    Add export metadata to the Export Info sheet.

    Includes information about the export operation:
    - Export timestamp
    - Source file path
    - Total rows exported
    - Exporter version

    Args:
        worksheet: openpyxl Worksheet object to format
        source_path: Path to source rawdata.xlsx file
        export_time: Export timestamp string
        row_count: Number of data rows exported
    """
    logger.info("Adding export metadata sheet")

    # Define metadata entries
    metadata = [
        ("Export Time", export_time),
        ("Source File", source_path),
        ("Total Rows", str(row_count)),
        ("Exported By", f"EDB Cutter Stackup Exporter v{EXPORTER_VERSION}")
    ]

    # Write metadata
    for row_idx, (key, value) in enumerate(metadata, start=1):
        key_cell = worksheet.cell(row=row_idx, column=1, value=key)
        key_cell.font = Font(bold=True)

        value_cell = worksheet.cell(row=row_idx, column=2, value=value)

    # Auto-adjust column widths
    _auto_adjust_column_widths(worksheet, ["Property", "Value"])

    logger.info("Export metadata sheet formatted successfully")


# Private helper functions for styling

def _apply_header_style(cell) -> None:
    """Apply header cell styling (bold white text on purple background)."""
    cell.font = Font(bold=True, color=HEADER_FONT_COLOR)
    cell.fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def _apply_alt_row_style(worksheet: Worksheet, row_idx: int, col_count: int) -> None:
    """Apply alternating row styling (subtle background color)."""
    for col_idx in range(1, col_count + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        cell.fill = PatternFill(start_color=ALT_ROW_FILL_COLOR, end_color=ALT_ROW_FILL_COLOR, fill_type="solid")


def _auto_adjust_column_widths(worksheet: Worksheet, headers: List[str]) -> None:
    """
    Auto-adjust column widths based on content.

    Sets minimum width based on header length, with additional padding.

    Args:
        worksheet: Worksheet to adjust
        headers: List of header strings for sizing reference
    """
    for col_idx, header in enumerate(headers, start=1):
        # Start with header length
        max_length = len(header)

        # Check data row lengths (sample first 100 rows for performance)
        for row_idx in range(2, min(102, worksheet.max_row + 1)):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        # Set column width with padding
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        column_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[column_letter].width = adjusted_width
