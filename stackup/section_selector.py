"""
Section selector module for stackup processing.
Handles section extraction from Excel and .sss file I/O operations.
"""
import json
import os
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from stackup.readers.excel_reader import extract_item_names_from_row8
from stackup.core.preprocessing import find_real_value


def extract_sections_from_excel(excel_file_path):
    """
    Extract section names from Excel file row 8.

    Args:
        excel_file_path (str): Path to Excel file

    Returns:
        list: List of section name strings

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        Exception: For other errors during extraction
    """
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"Excel file not found: {excel_file_path}")

    try:
        # Load Excel workbook
        wb = load_workbook(excel_file_path, data_only=True)
        ws = wb.worksheets[0]

        # Extract items from row 8
        items = extract_item_names_from_row8(ws)

        # Extract only the 'name' field from each item
        sections = [item['name'] for item in items if 'name' in item]

        return sections
    except Exception as e:
        raise Exception(f"Failed to extract sections from Excel: {str(e)}")


def generate_sss_filename(edb_folder_name):
    """
    Generate .sss filename with timestamp.

    Args:
        edb_folder_name (str): EDB folder name (e.g., "none_port_design.aedb")

    Returns:
        str: Filename in format "{edb_name}_sections_{timestamp}.sss"
    """
    # Remove .aedb extension if present
    base_name = edb_folder_name.replace('.aedb', '')

    # Generate timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    return f"{base_name}_sections_{timestamp}.sss"


def generate_layer_filename(edb_folder_name):
    """
    Generate layer .sss filename with timestamp.

    Args:
        edb_folder_name (str): EDB folder name

    Returns:
        str: Filename in format "{edb_name}_layers_{timestamp}.sss"
    """
    base_name = edb_folder_name.replace('.aedb', '')
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base_name}_layers_{timestamp}.sss"


def _find_spec_column(ws):
    """
    Find the SPEC column by searching for 'SPEC' keyword in row 8.

    Args:
        ws: Worksheet object

    Returns:
        int or None: Column number where 'SPEC' is found, or None if not found
    """
    # Search for 'SPEC' in row 8 (header row)
    for col in range(1, ws.max_column + 1):
        cell_value = find_real_value(ws, 8, col)
        if cell_value and 'SPEC' in str(cell_value).upper():
            return col
    return None


def _extract_spec_column_value(ws, row, spec_col):
    """
    Extract SPEC column value at given row.

    Args:
        ws: Worksheet object
        row (int): Row number
        spec_col (int): SPEC column number

    Returns:
        str or None: Spec value from SPEC column
    """
    if spec_col is None:
        return None
    spec_value = find_real_value(ws, row, spec_col)
    if spec_value:
        return str(spec_value).strip()
    return None


def extract_layer_data_for_section(excel_file_path, section_name):
    """
    Extract layer data for a specific section from Excel.

    Args:
        excel_file_path (str): Path to Excel file
        section_name (str): Name of section to extract layer data for

    Returns:
        list: Layer data for the section
        Example: [
            {
                'width': 50.0,
                'material': 'copper',
                'spec_name': 'SUS-TOP_1,3'
            },
            {
                'width': 12.0,
                'material': 'copper',
                'spec_name': 'EMI'
            }
        ]
    """
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"Excel file not found: {excel_file_path}")

    try:
        # Load workbook
        wb = load_workbook(excel_file_path, data_only=True)
        ws = wb.worksheets[0]

        # Get section items with column info
        items = extract_item_names_from_row8(ws)

        # Find the section's column
        section_col = None
        for item in items:
            if item.get('name') == section_name:
                section_col = item.get('col')
                break

        if section_col is None:
            raise ValueError(f"Section '{section_name}' not found in Excel")

        # Extract layer data from this column
        layer_data = []

        # Find SPEC column dynamically
        spec_col = _find_spec_column(ws)

        # Iterate through all rows starting from row 9 (after header)
        for current_row in range(9, ws.max_row + 1):
            # Read value from section column
            value = find_real_value(ws, current_row, section_col)

            # Skip if no value or value is not numeric
            if value is None:
                continue

            # Check if value is numeric (int or float)
            try:
                width = float(value)
            except (ValueError, TypeError):
                # Not a number, skip this row
                continue

            # Extract spec name from SPEC column at same row
            spec_name = _extract_spec_column_value(ws, current_row, spec_col)

            # Determine material based on spec name
            spec_lower = spec_name.lower() if spec_name else ''
            if 'space' in spec_lower or spec_lower == 's':
                material = 'air'
            else:
                material = 'copper'

            layer_data.append({
                'width': width,
                'material': material,
                'spec_name': spec_name
            })

        return layer_data

    except Exception as e:
        raise Exception(f"Failed to extract layer data for section '{section_name}': {str(e)}")


class SectionSelector:
    """
    Manages section selection workflow for stackup processing.

    Responsibilities:
    - Extract sections from Excel file
    - Save/load .sss configuration files
    - Validate cut-section mappings
    """

    def __init__(self, excel_file_path):
        """
        Initialize SectionSelector.

        Args:
            excel_file_path (str): Path to Excel file containing section data
        """
        self.excel_file = excel_file_path
        self._sections = None

    def extract_sections(self):
        """
        Extract section names from Excel file.
        Uses cached result if available.

        Returns:
            list: List of section name strings

        Raises:
            FileNotFoundError: If Excel file doesn't exist
            Exception: For other extraction errors
        """
        if self._sections is None:
            self._sections = extract_sections_from_excel(self.excel_file)
        return self._sections

    def save_section_mapping(self, cut_section_map, output_path):
        """
        Save cut-section mapping to .sss JSON file.

        Args:
            cut_section_map (dict): Mapping of cut IDs to section lists
                Example: {'cut_001': ['RIGID 5', 'C/N 1'], 'cut_002': [...]}
            output_path (str): Path to save .sss file

        Raises:
            ValueError: If cut_section_map is empty or invalid
            IOError: If file cannot be written
        """
        if not cut_section_map:
            raise ValueError("cut_section_map cannot be empty")

        # Get available sections
        sections = self.extract_sections()

        # Build .sss file structure
        sss_data = {
            "excel_file": str(Path(self.excel_file).as_posix()),
            "cut_section_mapping": cut_section_map,
            "available_sections": sections,
            "version": "1.0",
            "timestamp": datetime.now().isoformat()
        }

        # Save to file
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(sss_data, f, indent=2, ensure_ascii=False)
        except IOError as e:
            raise IOError(f"Failed to write .sss file: {str(e)}")

    def load_section_mapping(self, sss_file_path):
        """
        Load existing .sss configuration file.

        Args:
            sss_file_path (str): Path to .sss file

        Returns:
            dict: Loaded configuration data

        Raises:
            FileNotFoundError: If .sss file doesn't exist
            json.JSONDecodeError: If file is not valid JSON
        """
        if not os.path.exists(sss_file_path):
            raise FileNotFoundError(f".sss file not found: {sss_file_path}")

        try:
            with open(sss_file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            raise json.JSONDecodeError(f"Invalid .sss file format: {str(e)}", e.doc, e.pos)

    def validate_mapping(self, cut_section_map):
        """
        Validate cut-section mapping (1:1 mapping).

        Args:
            cut_section_map (dict): Mapping to validate
                Example: {'cut_001': 'RIGID 5', 'cut_002': 'C/N 1'}

        Returns:
            tuple: (is_valid: bool, error_messages: list)
        """
        errors = []

        # Check if mapping is empty
        if not cut_section_map:
            errors.append("Mapping is empty")
            return (False, errors)

        # Get available sections
        try:
            sections = self.extract_sections()
        except Exception as e:
            errors.append(f"Failed to extract sections for validation: {str(e)}")
            return (False, errors)

        # Validate each cut's section (1:1 mapping)
        for cut_id, selected_section in cut_section_map.items():
            if not selected_section:
                errors.append(f"Cut '{cut_id}' has no section selected")
                continue

            # Check if selected section is valid
            if selected_section not in sections:
                errors.append(f"Invalid section '{selected_section}' for cut '{cut_id}'")

        is_valid = len(errors) == 0
        return (is_valid, errors)

    def save_layer_data(self, cut_section_map, output_path):
        """
        Extract and save layer data for selected sections.

        Args:
            cut_section_map (dict): Mapping of cut IDs to sections
                Example: {'cut_001': 'RIGID 5', 'cut_002': 'C/N 1'}
            output_path (str): Path to save layer .sss file

        Raises:
            ValueError: If cut_section_map is empty
            IOError: If file cannot be written
        """
        if not cut_section_map:
            raise ValueError("cut_section_map cannot be empty")

        # Extract layer data for each selected section
        layer_data_by_cut = {}

        for cut_id, section_name in cut_section_map.items():
            try:
                layer_data = extract_layer_data_for_section(self.excel_file, section_name)
                layer_data_by_cut[cut_id] = {
                    'section': section_name,
                    'layers': layer_data
                }
            except Exception as e:
                # Log error but continue with other sections
                layer_data_by_cut[cut_id] = {
                    'section': section_name,
                    'layers': [],
                    'error': str(e)
                }

        # Build layer .sss file structure
        layer_sss_data = {
            "excel_file": str(Path(self.excel_file).as_posix()),
            "cut_layer_data": layer_data_by_cut,
            "version": "1.0",
            "timestamp": datetime.now().isoformat()
        }

        # Save to file
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(layer_sss_data, f, indent=2, ensure_ascii=False)
        except IOError as e:
            raise IOError(f"Failed to write layer .sss file: {str(e)}")
