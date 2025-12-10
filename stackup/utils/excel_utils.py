# excel_utils.py
"""
Excel utility functions for stackup processing.
Provides common Excel operations like sheet validation and text searching.
"""
from openpyxl import load_workbook
from util.logger_module import logger


class ExcelUtilities:
    """Utility functions for Excel file operations"""

    @staticmethod
    def check_sheet_exists(excel_file, sheet_name):
        """
        Check if a sheet exists in an Excel file

        Args:
            excel_file (str): Path to Excel file
            sheet_name (str): Name of sheet to check

        Returns:
            bool: True if sheet exists, False otherwise
        """
        try:
            wb = load_workbook(excel_file, data_only=True, read_only=True)
            exists = sheet_name in wb.sheetnames
            wb.close()
            return exists
        except Exception as e:
            logger.error(f"Error checking sheet '{sheet_name}': {e}")
            return False

    @staticmethod
    def find_text_cell_in_column(excel_file, sheet_name, column):
        """
        Find first text cell in specified column

        Args:
            excel_file (str): Path to Excel file
            sheet_name (str): Sheet name
            column (str): Column letter (e.g., 'C')

        Returns:
            str: Cell address like 'C4' or None if not found
        """
        try:
            wb = load_workbook(excel_file, data_only=True, read_only=True)
            ws = wb[sheet_name]

            # Search for first non-empty cell in column
            for row in range(1, ws.max_row + 1):
                cell = ws[f'{column}{row}']
                if cell.value and isinstance(cell.value, str):
                    wb.close()
                    return f'{column}{row}'

            wb.close()
            return None
        except Exception as e:
            logger.error(f"Error finding text in column {column}: {e}")
            return None

    @staticmethod
    def get_column_letter(column_number):
        """
        Convert column number to letter (1 -> 'A', 2 -> 'B', etc.)

        Args:
            column_number (int): Column number (1-indexed)

        Returns:
            str: Column letter
        """
        from openpyxl.utils import get_column_letter
        return get_column_letter(column_number)
