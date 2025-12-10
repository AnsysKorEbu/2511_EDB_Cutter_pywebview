# excel_reader.py
"""
Excel file reader for FPCB cross-section material properties.
Handles complex material specification patterns like (3-1,4-1L) for position-based matching.
"""
import re
import os
import sys
from openpyxl import load_workbook
from util.logger_module import logger
from stackup.core.config import get_height_column
from stackup.core.preprocessing import check_file_exists, find_real_value, sanitize_name

# Constants for better readability
MAIN_MATERIALS_KEYWORD = 'MAIN MATERIALS'
SUBSIDIARY_MATERIALS_KEYWORD = 'SUBSIDIARY MATERIALS'
SPEC_SEARCH_COLUMNS = range(18, 25)  # Columns to search for specifications
PSR_SEARCH_COLUMNS = range(6, 36)    # Wider range for PSR materials
DK_DF_PATTERN = r'\(([\d\s\.]*\s*/\s*[\d\s\.]*)\)\s*10GHz?'  # Pattern to match Dk/Df values
POSITION_PATTERN = r'(\d+)-(\d+)'     # Pattern to match (3-1,4-1L) style positions
THICKNESS_PATTERN = r'(\d+(?:\.\d+)?)\s*(?:μm|um|micron)?'  # Pattern to match thickness values


def _find_section_start_row(ws, keyword):
    """
    Find the starting row of a section by searching for a keyword.

    Args:
        ws: Openpyxl worksheet object
        keyword (str): Keyword to search for (e.g., 'MAIN MATERIALS')

    Returns:
        int or None: Row number where the section starts, or None if not found
    """
    for row in range(1, ws.max_row + 1):
        cell_value = find_real_value(ws, row, 2)
        if cell_value and keyword in str(cell_value):
            return row
    return None


def _process_position_pattern(material_type, subtype_key, spec_value, result_dict):
    """
    Process patterns like (3-1,4-1L) and create separate entries.

    Args:
        material_type (str): Base material type (e.g., 'COVERLAY')
        subtype_key (str): Subtype with pattern (e.g., '(3-1,4-1L)')
        spec_value (str): Specification value
        result_dict (dict): Dictionary to add processed entries to

    Returns:
        bool: True if pattern was processed, False otherwise
    """
    if '(' in subtype_key and 'L)' in subtype_key:
        # Extract pattern like 3-1, 4-1 from (3-1,4-1L)
        matches = re.findall(POSITION_PATTERN, subtype_key)
        if matches:
            for layer_num, position in matches:
                # Create key like "COVERLAY (3L-1)" for 3rd layer, 1st position
                layer_key = f"{material_type} ({layer_num}L-{position})"
                result_dict[layer_key] = spec_value
            return True
    return False


def _is_valid_specification(value):
    """Check if a value contains valid specification indicators."""
    value_str = str(value).lower()
    return 'ghz' in value_str or '/' in value_str or '(' in value_str


def _find_and_store_specification(ws, current_row, material_type, material_subtype, search_columns, result_dict, is_subsidiary=False):
    """
    Find material specification and store it in result dictionary.
    Handles both main and subsidiary materials with their specific search patterns.
    """
    spec_value = None

    # Search for specifications in designated columns
    for col in search_columns:
        spec_candidate = find_real_value(ws, current_row, col)
        if spec_candidate and str(spec_candidate).strip() and str(spec_candidate).strip() != material_type:
            spec_candidate = str(spec_candidate).strip()

            if is_subsidiary:
                # For PSR, look for complete specification with parentheses
                if material_type == 'PSR':
                    if '(' in spec_candidate and ')' in spec_candidate:
                        spec_value = spec_candidate
                        break
                else:
                    # For other subsidiary materials, any valid specification is acceptable
                    spec_value = spec_candidate
                    break
            else:
                # For main materials, check if it's a valid specification
                if _is_valid_specification(spec_candidate):
                    spec_value = spec_candidate
                    break

    # Store the specification if found
    if spec_value:
        if is_subsidiary:
            result_dict[material_type] = spec_value
        else:
            # Handle main materials with subtypes and position patterns
            if material_subtype and str(material_subtype).strip():
                subtype_key = str(material_subtype).strip()

                # Try to process position pattern first
                if _process_position_pattern(material_type, subtype_key, spec_value, result_dict):
                    return

                # Fall back to regular key format
                combined_key = f"{material_type}/{subtype_key}"
                result_dict[combined_key] = spec_value
            else:
                # No subtype, use material type as key
                result_dict[material_type] = spec_value


def _process_materials_sections(ws, main_start_row, subsidiary_start_row, main_result, subsidiary_result):
    """
    Process both main and subsidiary materials sections.
    Consolidates the processing logic for both types of materials.
    """
    # Process main materials section
    if main_start_row:
        current_row = main_start_row + 3  # Skip header rows

        while current_row < ws.max_row:
            # Stop if we reach subsidiary materials section
            if subsidiary_start_row and current_row >= subsidiary_start_row:
                break

            # Get material information from current row
            material_type = find_real_value(ws, current_row, 2)
            material_subtype = find_real_value(ws, current_row, 8)  # Film/Adhesive/Copper/Polyimide

            # Skip empty rows
            if not material_type:
                current_row += 1
                continue

            material_type = str(material_type).strip()

            # Find and store specification
            _find_and_store_specification(
                ws, current_row, material_type, material_subtype,
                SPEC_SEARCH_COLUMNS, main_result, is_subsidiary=False
            )

            current_row += 1

    # Process subsidiary materials section
    if subsidiary_start_row:
        current_row = subsidiary_start_row + 3  # Skip header rows
        max_rows_to_process = 50  # Avoid infinite loops

        while current_row < ws.max_row and current_row < subsidiary_start_row + max_rows_to_process:
            material_type = find_real_value(ws, current_row, 2)

            if not material_type:
                current_row += 1
                continue

            material_type = str(material_type).strip()

            # Find and store specification for subsidiary materials
            _find_and_store_specification(
                ws, current_row, material_type, None,
                PSR_SEARCH_COLUMNS, subsidiary_result, is_subsidiary=True
            )

            current_row += 1


def _extract_clean_dk_df_value(spec_value):
    """
    Extract and clean Dk/Df value from specification string.

    Args:
        spec_value (str): Raw specification string

    Returns:
        str or None: Cleaned Dk/Df value like '3.17/0.023(10GHz)', or None
    """
    match = re.search(DK_DF_PATTERN, spec_value, re.IGNORECASE)
    if match:
        dk_df_part = match.group(1).strip()
        # Check if it's empty like "  /  "
        if dk_df_part.replace('/', '').strip() == '':
            return None
        else:
            # Remove extra spaces and format as dk/df(10GHz)
            cleaned_dk_df = dk_df_part.replace(' ', '')
            return f"{cleaned_dk_df}(10GHz)"
    return None


def _process_material_key(key, dk_df_value, result_list):
    """
    Process a material key and add appropriate entries to result list.

    Args:
        key (str): Material key from raw dictionary
        dk_df_value (str or None): Cleaned Dk/Df value
        result_list (list): List to append results to
    """
    if '(' in key and 'L)' in key:
        _process_layer_pattern_key(key, dk_df_value, result_list)
    elif 'prepreg' in key.lower():
        _process_prepreg_key(key, dk_df_value, result_list)
    else:
        # Simple material without layer pattern
        result_list.append({
            'layer': None,
            'material': key.lower(),
            'dk_df': dk_df_value
        })


def _process_layer_pattern_key(key, dk_df_value, result_list):
    """
    Process keys with layer patterns like (1,6L) or (3L-1).

    Args:
        key (str): Material key with layer pattern
        dk_df_value (str or None): Dk/Df value
        result_list (list): List to append results to
    """
    # Check for position-based pattern first (3L-1)
    position_pattern = r'(\w+)\s*\((\d+)L-(\d+)\)'
    position_match = re.search(position_pattern, key)

    if position_match:
        # Handle position-based pattern like 'COVERLAY (3L-1)'
        material_type = position_match.group(1).lower()
        result_list.append({
            'layer': key,  # Keep full key as layer identifier for position matching
            'material': material_type,
            'dk_df': dk_df_value
        })
        return

    # Handle regular layer patterns like 'FCCL (1,6L)/Copper'
    parts = key.split('/')
    if len(parts) == 2:
        material_base = parts[0].strip()  # 'FCCL (1,6L)'
        subtype = parts[1].strip().lower()  # 'copper'

        # Extract layer numbers from parentheses
        layer_pattern = r'\(([^)]+)\)'
        layer_match = re.search(layer_pattern, material_base)

        if layer_match:
            layer_info = layer_match.group(1).rstrip('L')  # Remove trailing 'L'
            layer_numbers = [num.strip().rstrip('L') for num in layer_info.split(',')]

            for layer_num in layer_numbers:
                if layer_num:  # Skip empty strings
                    # Copper materials don't have Dk/Df values
                    final_dk_df = None if subtype.lower() == 'copper' else dk_df_value
                    result_list.append({
                        'layer': f"{layer_num}LAYER",
                        'material': subtype,
                        'dk_df': final_dk_df
                    })


def _process_prepreg_key(key, dk_df_value, result_list):
    """
    Process PREPREG keys with patterns like (12,23,45,56) for inter-layer adhesives.

    Args:
        key (str): PREPREG key with layer numbers
        dk_df_value (str or None): Dk/Df value
        result_list (list): List to append results to
    """
    # Extract numbers from parentheses
    layer_pattern = r'\(([^)]+)\)'
    layer_match = re.search(layer_pattern, key)

    if layer_match:
        numbers_str = layer_match.group(1)  # '12,23,45,56' or '34'
        numbers = [num.strip() for num in numbers_str.split(',')]

        for num in numbers:
            if len(num) == 2:  # Like '12', '23', '45', '56' (between layers 1-2, 2-3, etc.)
                start_layer = int(num[0])
                end_layer = int(num[1])
                middle_layer = (start_layer + end_layer) / 2
                result_list.append({
                    'layer': f"{middle_layer}LAYER",
                    'material': 'adhesive',
                    'dk_df': dk_df_value
                })
            elif len(num) == 1:  # Like '3', '4' (above layer 3, above layer 4)
                layer_num = int(num)
                if layer_num > 1:
                    middle_layer = layer_num - 0.5
                    result_list.append({
                        'layer': f"{middle_layer}LAYER",
                        'material': 'adhesive',
                        'dk_df': dk_df_value
                    })


def extract_item_names_from_row8(ws, target_row=None):
    """
    Extract header names from merged cells in the row containing 'layer'.
    This function searches for the first row with 'layer' text and processes merged cells.

    Args:
        ws: Openpyxl worksheet object
        target_row (int, optional): Row number to extract headers from (for backward compatibility)

    Returns:
        list: List of dictionaries containing header information
        Example: [{'name': 'LAYER', 'col': 2}, {'name': 'SPEC.', 'col': 5}]

    Note:
        This function will be used for future column mapping functionality.
    """
    if target_row is None:
        # Find first row containing 'layer' (case insensitive)
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    if str(cell_value).lower().strip() == 'layer':
                        target_row = row
                    break
            if target_row:
                break

        if target_row is None:
            # Fallback to row 8 if 'layer' not found
            target_row = 8

    items = []
    for merged_range in ws.merged_cells.ranges:
        if target_row >= merged_range.min_row and target_row <= merged_range.max_row:
            value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            if value:
                clean_name = str(value).replace('\n', ' ').strip()
                items.append({
                    'name': clean_name,
                    'col': merged_range.min_col
                })

    # Sort by column position
    items.sort(key=lambda x: x['col'])
    return items[4:]


def extract_materials_specifications(ws):
    """
    Extract material specifications from Excel worksheet.

    This function processes two main sections:
    1. A. MAIN MATERIALS section - Contains FCCL, COVERLAY, PREPREG specifications
    2. B. SUBSIDIARY MATERIALS section - Contains PSR, INK, etc.

    Special handling for patterns like (3-1,4-1L):
    - Creates separate entries for each layer-position combination
    - Example: COVERLAY(3-1,4-1L) → COVERLAY(3L-1), COVERLAY(4L-1)

    Args:
        ws: Openpyxl worksheet object

    Returns:
        dict: Dictionary with 'main_materials' and 'subsidiary_materials' keys
    """
    result = {
        'main_materials': {},
        'subsidiary_materials': {}
    }

    # Locate the materials sections in the worksheet
    main_materials_start_row = _find_section_start_row(ws, MAIN_MATERIALS_KEYWORD)
    subsidiary_materials_start_row = _find_section_start_row(ws, SUBSIDIARY_MATERIALS_KEYWORD)

    # Process both sections using consolidated function
    _process_materials_sections(
        ws, main_materials_start_row, subsidiary_materials_start_row,
        result['main_materials'], result['subsidiary_materials']
    )

    return result


def postprocess_materials_dict(dk_df_dict):
    """
    Convert raw materials dictionary to structured list with clean Dk/Df values.

    This function:
    1. Extracts and cleans Dk/Df values from specification strings
    2. Expands layer patterns like (1,6L) into separate entries
    3. Handles position patterns like (3L-1) for specific layer positions
    4. Processes PREPREG patterns for inter-layer adhesives

    Args:
        dk_df_dict (dict): Raw materials from extract_materials_specifications

    Returns:
        list: Processed material entries
        Example:
        [
            {'layer': '4-1LAYER', 'material': 'film', 'dk_df': '3.17/0.023(10GHz)'},
            {'layer': '2LAYER', 'material': 'copper', 'dk_df': None},
            {'layer': None, 'material': 'psr', 'dk_df': '3.6/0.027(10GHz)'}
        ]
    """
    result = []

    # Process main materials
    for key, value in dk_df_dict.get('main_materials', {}).items():
        cleaned_dk_df = _extract_clean_dk_df_value(value)
        _process_material_key(key, cleaned_dk_df, result)

    # Process subsidiary materials (mainly PSR)
    for key, value in dk_df_dict.get('subsidiary_materials', {}).items():
        cleaned_dk_df = _extract_clean_dk_df_value(value)
        if cleaned_dk_df and key.upper() == 'PSR':  # Only add PSR with valid dk_df
            result.append({
                'layer': None,
                'material': key.lower(),
                'dk_df': cleaned_dk_df
            })

    return result


def _check_compatibility(entry_material, mat_material, entry_layer, mat_layer):
    """
    Check if materials and layers are compatible for matching.

    Args:
        entry_material (str): Entry material name
        mat_material (str): Material from data
        entry_layer (str): Entry layer identifier
        mat_layer (str or None): Material layer identifier

    Returns:
        bool: True if both materials and layers are compatible
    """
    # Check material compatibility
    material_match = (mat_material in entry_material or
                     entry_material in mat_material or
                     'film' in mat_material and 'film' in entry_material or
                     'adhesive' in mat_material and 'adhesive' in entry_material or
                     mat_material == 'psr' and 'psr' in entry_material)

    # Check layer compatibility
    if mat_layer is None:
        layer_match = True  # Materials without specific layers match all layers
    elif mat_layer == entry_layer:
        layer_match = True
    else:
        # Try extracting numbers for comparison
        try:
            entry_num = entry_layer.replace('LAYER', '').replace('layer', '')
            mat_num = mat_layer.replace('LAYER', '').replace('layer', '')
            layer_match = entry_num == mat_num
        except:
            layer_match = False

    return material_match and layer_match


def match_dk_df_for_entry(entry_layer, entry_material, dk_df_data, layer_material_info=None, layer_material_positions=None, current_row=None):
    """
    Find matching Dk/Df value for a layer and material using position-based or standard matching.

    Args:
        entry_layer (str): Layer identifier (e.g., "4LAYER")
        entry_material (str): Material name (e.g., "C_L_Film")
        dk_df_data (list): Processed materials data
        layer_material_info (list): Full layer material info (unused, kept for compatibility)
        layer_material_positions (dict): Pre-calculated position info
        current_row (int): Row number for position lookup

    Returns:
        str or None: Matching Dk/Df value or None if no match
    """
    if not entry_layer or not entry_material or not dk_df_data:
        return None

    entry_layer_str = str(entry_layer).strip()
    entry_material_str = str(entry_material).lower().strip()

    # Strategy 1: Position-based matching (primary method)
    if layer_material_positions and current_row:
        try:
            layer_num = int(entry_layer_str.replace('LAYER', '').replace('layer', ''))

            if layer_num in layer_material_positions:
                position_key = f"{entry_material_str}_{current_row}"
                material_position = layer_material_positions[layer_num].get(position_key)

                if material_position:
                    expected_key = f"{layer_num}-{material_position}LAYER"

                    for mat_entry in dk_df_data:
                        mat_layer = mat_entry.get('layer')
                        mat_material = mat_entry.get('material', '').lower()
                        mat_dk_df = mat_entry.get('dk_df')

                        if (mat_dk_df and mat_layer == expected_key and
                            _check_compatibility(entry_material_str, mat_material, entry_layer_str, mat_layer)):
                            return mat_dk_df
        except (ValueError, AttributeError):
            pass

    # Strategy 2: Standard layer-material matching (fallback)
    for mat_entry in dk_df_data:
        mat_layer = mat_entry.get('layer')
        mat_material = mat_entry.get('material', '').lower()
        mat_dk_df = mat_entry.get('dk_df')

        if mat_dk_df and _check_compatibility(entry_material_str, mat_material, entry_layer_str, mat_layer):
            return mat_dk_df

    return None


def calculate_adhesive_air_gap(ws, current_row, height):
    """
    Calculate air gap height for adhesive materials by comparing heights.

    Args:
        ws: Openpyxl worksheet object
        current_row (int): Current row number
        height (float): Current material height

    Returns:
        float or None: Air gap height, or None if no gap exists
    """
    # Search for other material heights in the same row
    other_heights = []

    for col in range(16, 130):
        if col == 129:  # Skip specific column
            continue

        cell_value = find_real_value(ws, current_row, col)
        if cell_value is not None:
            try:
                numeric_value = float(cell_value)
                other_heights.append(numeric_value)
            except (ValueError, TypeError):
                continue

    # Calculate air gap if taller materials exist
    if other_heights:
        max_other_height = max(other_heights)
        if max_other_height > height:
            return max_other_height - height

    return None


def _calculate_material_positions(layer_material_info):
    """Calculate material positions within each layer for position-based matching."""
    layer_material_positions = {}
    for info in layer_material_info:
        layer_num = info.get('layer')
        material = info.get('material', '').lower()

        if layer_num not in layer_material_positions:
            layer_material_positions[layer_num] = {}

        # Count film/adhesive materials to determine their positions within each layer
        for material_type in ['film', 'adhesive']:
            if material_type in material:
                if material_type not in layer_material_positions[layer_num]:
                    layer_material_positions[layer_num][material_type] = 0
                layer_material_positions[layer_num][material_type] += 1

                # Store position for this specific material instance
                position_key = f"{material}_{info.get('row')}"
                layer_material_positions[layer_num][position_key] = layer_material_positions[layer_num][material_type]

    return layer_material_positions


def _process_material_rows(ws, dk_df_dict, layer_material_info, layer_material_positions):
    """Process each material row in the worksheet and create material entries."""
    current_row = 9
    result = []

    while True:
        check_value = ws.cell(row=current_row, column=5).value
        if check_value is None:
            break

        # Extract basic material information
        key_name = find_real_value(ws, current_row, 2)
        material_name = find_real_value(ws, current_row, 5)
        height = find_real_value(ws, current_row, get_height_column())

        if height is None:
            current_row += 1
            continue

        # Create material entry
        entry = _create_material_entry(
            ws, current_row, key_name, material_name, height,
            dk_df_dict, layer_material_info, layer_material_positions
        )
        result.append(entry)

        # Handle air gaps for adhesive materials
        if material_name and 'adhesive' in material_name.strip().lower():
            air_gap_height = calculate_adhesive_air_gap(ws, current_row, height)
            if air_gap_height is not None:
                air_entry = {
                    'layer': key_name, 'row': current_row, 'material': 'air',
                    'CU_foil': None, 'Dk/Df': '1.006/0(10GHz)', 'height': air_gap_height
                }
                result.append(air_entry)

        current_row += 1

    return result


def _create_material_entry(ws, current_row, key_name, material_name, height,
                          dk_df_dict, layer_material_info, layer_material_positions):
    """Create a single material entry with Dk/Df matching."""
    cu_foil = material_name if material_name and 'Copper' in str(material_name) else None

    # Find matching Dk/Df value
    matched_dk_df = match_dk_df_for_entry(
        key_name, sanitize_name(material_name), dk_df_dict,
        layer_material_info, layer_material_positions, current_row
    )

    return {
        'layer': key_name,
        'row': current_row,
        'material': sanitize_name(material_name),
        'CU_foil': cu_foil,
        'Dk/Df': matched_dk_df,
        'height': height
    }


def read_material_properties(filename):
    """
    Read material properties from Excel file and match with Dk/Df values.

    Args:
        filename (str): Path to Excel file

    Returns:
        list: Material properties with Dk/Df values and heights
    """
    if not check_file_exists(filename):
        sys.exit(1)

    wb = load_workbook(filename, data_only=True)
    ws = wb.worksheets[0]

    # items = extract_item_names_from_row8(ws)


    # Get materials specifications and layer info
    dk_df_dict = extract_materials_specifications(ws)
    dk_df_dict = postprocess_materials_dict(dk_df_dict)
    layer_material_info = read_layer_material(filename)
    layer_material_positions = _calculate_material_positions(layer_material_info)

    # Process material rows
    result = _process_material_rows(ws, dk_df_dict, layer_material_info, layer_material_positions)

    # Apply thickness-based Dk/Df combinations
    dk_df_by_thickness = create_dk_df_by_thickness(ws)
    result = apply_dk_df_from_thickness_combinations(result, dk_df_by_thickness)

    return result


def read_layer_material(filename):
    """
    Read layer material information from Excel file.

    Args:
        filename (str): Path to Excel file

    Returns:
        list: Layer material data with layer, row, and material info
    """
    if not check_file_exists(filename):
        sys.exit(1)

    wb = load_workbook(filename, data_only=True)
    ws = wb.worksheets[0]
    current_row = 9
    result = []

    # Find first EMI material
    emi_found = False
    emi_count = 0

    while True:
        # Check if current row has data (column 4)
        check_value = ws.cell(row=current_row, column=5).value
        if check_value is None:
            break

        # Extract data from columns
        key_name = find_real_value(ws, current_row, 2)  # layer (column 2)
        material_name = find_real_value(ws, current_row, 5)  # material (column 4)

        # Check if current material is EMI
        if material_name and 'EMI' in str(material_name).upper():
            emi_count += 1
            if not emi_found:
                emi_found = True  # Found first EMI
            elif emi_count > 1:
                break  # Stop when we encounter the second EMI

        # Only process if we found the first EMI
        if not emi_found:
            current_row += 1
            continue

        if key_name is not None and material_name is not None:
            # Convert layer name to integer if possible
            layer_value = key_name
            if isinstance(key_name, str) and 'LAYER' in key_name.upper():
                try:
                    layer_value = int(key_name.upper().replace('LAYER', ''))
                except ValueError:
                    layer_value = key_name
            elif isinstance(key_name, str) and key_name.isdigit():
                layer_value = int(key_name)

            # Create entry dictionary with layer, row, and material
            entry = {
                'layer': layer_value,
                'material': sanitize_name(material_name),
                'row': current_row
            }
            result.append(entry)

        current_row += 1

    return result


def _check_material_combination_match(adhesive_material, film_material, dk_df_entry):
    """
    Check if adhesive and film materials match a dk_df_by_thickness entry.

    Args:
        adhesive_material (str): Adhesive material name
        film_material (str): Film material name
        dk_df_entry (dict): Entry from dk_df_by_thickness with 'layer' and 'dk/df' keys

    Returns:
        bool: True if the combination matches the dk_df_entry
    """
    if 'layer' not in dk_df_entry or not isinstance(dk_df_entry['layer'], dict):
        return False

    layer_dict = dk_df_entry['layer']

    # Check for adhesive and film components in the layer dictionary
    has_adhesive = any('adhesive' in subtype.lower() for subtype in layer_dict.keys())
    has_film = any('film' in subtype.lower() for subtype in layer_dict.keys())

    # Basic requirement: must have both adhesive and film
    if not (has_adhesive and has_film):
        return False

    # Enhanced matching: check if material names suggest a specific type
    # For C_L_* materials, prefer COVERLAY matches
    adhesive_clean = adhesive_material.lower().replace('_', '').replace('m', '')  # Handle typos like "adhesivem"
    film_clean = film_material.lower().replace('_', '')

    # If materials are C_L_* type, prefer COVERLAY entries
    if 'c' in adhesive_clean and 'l' in adhesive_clean and 'c' in film_clean and 'l' in film_clean:
        return True  # This is a COVERLAY-type combination

    # For other combinations, accept any adhesive/film combination
    return True


def _find_best_thickness_match(adhesive_material, film_material, layer_info, dk_df_by_thickness):
    """Find the best Dk/Df match from thickness data, preferring layer-specific matches."""
    best_match = None
    best_dk_df = None

    for material_type, dk_df_entry in dk_df_by_thickness.items():
        if _check_material_combination_match(adhesive_material, film_material, dk_df_entry):
            dk_df_value = dk_df_entry.get('dk/df')
            if dk_df_value:
                # Check for layer-specific matches (e.g., (3-1,4-1L) for 3LAYER)
                if layer_info and '3' in layer_info and ('3' in material_type or '3-1' in material_type):
                    return dk_df_value  # Prefer layer-specific matches
                elif not best_match:  # Store as fallback
                    best_match = material_type
                    best_dk_df = dk_df_value

    return best_dk_df


def _process_consecutive_materials(current_entry, next_entry, dk_df_by_thickness):
    """Process a pair of consecutive materials to apply Dk/Df if they form adhesive/film combination."""
    if (current_entry.get('Dk/Df') is not None or next_entry.get('Dk/Df') is not None or
        not current_entry.get('material') or not next_entry.get('material')):
        return False

    current_material = current_entry['material'].lower()
    next_material = next_entry['material'].lower()
    layer_info = current_entry.get('layer', '')

    # Check for adhesive->film or film->adhesive combinations
    is_adhesive_film = 'adhesive' in current_material and 'film' in next_material
    is_film_adhesive = 'film' in current_material and 'adhesive' in next_material

    if is_adhesive_film:
        best_dk_df = _find_best_thickness_match(current_material, next_material, layer_info, dk_df_by_thickness)
    elif is_film_adhesive:
        best_dk_df = _find_best_thickness_match(next_material, current_material, layer_info, dk_df_by_thickness)
    else:
        return False

    # Apply the best match found
    if best_dk_df:
        current_entry['Dk/Df'] = best_dk_df
        next_entry['Dk/Df'] = best_dk_df
        return True

    return False


def apply_dk_df_from_thickness_combinations(result, dk_df_by_thickness):
    """
    Apply Dk/Df values from dk_df_by_thickness for consecutive adhesive/film combinations.

    Args:
        result (list): List of material entries
        dk_df_by_thickness (dict): Thickness-based material data

    Returns:
        list: Updated result list with applied Dk/Df values
    """
    if not result or not dk_df_by_thickness:
        return result

    # Process consecutive material pairs
    for i in range(len(result) - 1):
        _process_consecutive_materials(result[i], result[i + 1], dk_df_by_thickness)

    return result


def create_dk_df_by_thickness(ws):
    """
    Create structured dictionary with material thickness and Dk/Df information.
    Extracts data from MAIN MATERIALS section columns 2, 8, 13, 18.

    Args:
        ws: Openpyxl worksheet object

    Returns:
        dict: Dictionary structured as {material: {layer: {subtype: thickness}, 'dk/df': value}}
        Example: {'COVERLAY': {'layer': {'Film': 12.5, 'Adhesive Polyimide': 15}, 'dk/df': '(3.26 / 0.019) 10GHz'}}
    """
    dk_df_by_thickness = {}

    # Find MAIN MATERIALS section
    main_materials_start_row = _find_section_start_row(ws, MAIN_MATERIALS_KEYWORD)
    subsidiary_materials_start_row = _find_section_start_row(ws, SUBSIDIARY_MATERIALS_KEYWORD)

    if not main_materials_start_row:
        return dk_df_by_thickness

    # Define columns to extract data from (2, 8, 13, 18)
    data_columns = [2, 8, 13, 18]
    current_row = main_materials_start_row + 3  # Skip header rows

    while current_row < ws.max_row:
        # Stop if we reach subsidiary materials section
        if subsidiary_materials_start_row and current_row >= subsidiary_materials_start_row:
            break

        # Get material type from column 2
        material_type = find_real_value(ws, current_row, 2)
        if not material_type:
            current_row += 1
            continue

        material_type = str(material_type).strip()

        # Get material subtype from column 8 to use as thickness key
        material_subtype = find_real_value(ws, current_row, 8)
        subtype_str = str(material_subtype).strip() if material_subtype else None

        # Get thickness value from column 13 (THICK)
        thickness_value = find_real_value(ws, current_row, 13)

        # Search for specifications in columns 18 and beyond for Dk/Df values
        spec_value = None
        for col in range(18, 25):
            spec_candidate = find_real_value(ws, current_row, col)
            if spec_candidate and _is_valid_specification(spec_candidate):
                spec_value = str(spec_candidate).strip()
                break

        # Process if we have subtype and thickness, or valid specification
        if (subtype_str and thickness_value is not None) or spec_value:
            # Initialize material entry if not exists
            if material_type not in dk_df_by_thickness:
                dk_df_by_thickness[material_type] = {
                    'layer': {},
                    'dk/df': None
                }

            # Add thickness information if available
            if subtype_str and thickness_value is not None:
                try:
                    dk_df_by_thickness[material_type]['layer'][subtype_str] = float(thickness_value)
                except (ValueError, TypeError):
                    pass

            # Add Dk/Df information if available
            if spec_value:
                dk_df_value = _extract_clean_dk_df_value(spec_value)
                if dk_df_value:
                    dk_df_by_thickness[material_type]['dk/df'] = dk_df_value

        current_row += 1

    return dk_df_by_thickness


def _extract_thickness_info_with_subtype(spec_value, subtype_str):
    """
    Extract thickness information using material subtype as key.

    Args:
        spec_value (str): Specification string containing thickness info
        subtype_str (str): Material subtype to use as thickness key (e.g., "Adhesive Polyimide")

    Returns:
        dict: Dictionary with subtype as key and thickness as value
        Example: {'Adhesive Polyimide': 15.0, 'Film': 12.5}
    """
    thickness_info = {}

    if not subtype_str:
        return thickness_info

    # Extract thickness values from specification
    thickness_matches = re.findall(THICKNESS_PATTERN, spec_value)

    if thickness_matches:
        # Use the first thickness value found
        thickness_value = float(thickness_matches[0])

        # Use subtype_str as the key (e.g., "Adhesive Polyimide", "Film", "Copper")
        thickness_info[subtype_str] = thickness_value

    return thickness_info


def _extract_thickness_info(spec_value):
    """
    Extract thickness information for film and adhesive from specification string.

    Args:
        spec_value (str): Specification string containing thickness info

    Returns:
        dict: Dictionary with 'film' and 'adhesive' thickness values
        Example: {'film': 12.5, 'adhesive': 15.0}
    """
    thickness_info = {'film': None, 'adhesive': None}

    # Convert to lowercase for easier matching
    spec_lower = spec_value.lower()

    # Look for thickness patterns with context
    # Common patterns: "Film 12.5μm", "Adhesive 15μm", "12.5/15" format

    # Pattern 1: Explicit film/adhesive mentions
    film_match = re.search(r'film[:\s]*(\d+(?:\.\d+)?)', spec_lower)
    if film_match:
        thickness_info['film'] = float(film_match.group(1))

    adhesive_match = re.search(r'adhesive[:\s]*(\d+(?:\.\d+)?)', spec_lower)
    if adhesive_match:
        thickness_info['adhesive'] = float(adhesive_match.group(1))

    # Pattern 2: Slash-separated format (film/adhesive)
    slash_pattern = re.search(r'(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)', spec_value)
    if slash_pattern and not thickness_info['film'] and not thickness_info['adhesive']:
        # Only use if we haven't found explicit film/adhesive values
        thickness_info['film'] = float(slash_pattern.group(1))
        thickness_info['adhesive'] = float(slash_pattern.group(2))

    # Pattern 3: Multiple thickness values mentioned
    thickness_matches = re.findall(THICKNESS_PATTERN, spec_value)
    if thickness_matches and not any(thickness_info.values()):
        # If we found thickness values but couldn't categorize them
        for i, thickness in enumerate(thickness_matches):
            if i == 0 and not thickness_info['film']:
                thickness_info['film'] = float(thickness)
            elif i == 1 and not thickness_info['adhesive']:
                thickness_info['adhesive'] = float(thickness)

    return thickness_info


if __name__ == "__main__":
    print("Testing excel_reader functions...")

    try:
        from stackup.utils.config_functions import get_excel_file_rawdata
        excel_file_rawdata = get_excel_file_rawdata()

        if not excel_file_rawdata:
            print("Error: Could not find rawdata Excel file")
            sys.exit(1)

        layer_data = read_material_properties(excel_file_rawdata)
        print(f"Material properties data: {layer_data}")
        print("")

        layer_material_info = read_layer_material(excel_file_rawdata)
        print(f"Layer material info: {layer_material_info}")
        print("")

    except Exception as e:
        print(f"Error occurred: {e}")
        import traceback
        traceback.print_exc()
