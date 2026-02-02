"""
Data preprocessing functions for stackup processing.
Standalone version without external dependencies.
"""

import os
from openpyxl import load_workbook
from .logger import logger


def find_real_value(ws, row, col):
    """
    Find real value in worksheet, handling merged cells.

    Args:
        ws: Openpyxl worksheet object
        row (int): Row number
        col (int): Column number

    Returns:
        Cell value (handles merged cells)
    """
    for merged_range in ws.merged_cells.ranges:
        if (row, col) in merged_range.cells:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return ws.cell(row=row, column=col).value


def sanitize_name(net):
    """
    Sanitize material/net names for consistency.

    Args:
        net (str): Material or net name

    Returns:
        str: Sanitized name
    """
    if net is None:
        return net
    net = net.replace('-', 'm')
    net = net.replace('.', 'p')
    net = net.replace(' ', '_')
    net = net.replace('/', '_')
    return net


def check_file_exists(filename):
    """
    Check if a file exists and log error if not.

    Args:
        filename (str): Path to the file to check

    Returns:
        bool: True if file exists, False otherwise
    """
    if os.path.isfile(filename):
        return True

    # Log the error
    file_name = os.path.basename(filename)
    logger.error(f"Required file not found: {filename}")
    return False


def find_center_layer(entries):
    """
    Find the center layer row in stackup entries.

    Args:
        entries (list): List of layer entries

    Returns:
        int or None: Row number of center layer
    """
    # Extract layer numbers, converting strings like '1LAYER' to integers
    layers = []
    for e in entries:
        if e['layer'] is not None:
            layer_val = e['layer']
            # Convert string layer names to integers
            if isinstance(layer_val, str):
                try:
                    # Extract number from strings like '1LAYER'
                    layer_num = int(layer_val.upper().replace('LAYER', '').strip())
                    layers.append(layer_num)
                except ValueError:
                    continue
            elif isinstance(layer_val, (int, float)):
                layers.append(int(layer_val))

    if not layers:
        return None

    mid = max(layers) // 2
    for i, entry in enumerate(entries):
        if i == 0 or i >= len(entries) - 1:
            continue

        prev_layer = entries[i - 1]['layer']
        next_layer = entries[i + 1]['layer']

        # Convert to int for comparison
        try:
            if isinstance(prev_layer, str):
                prev_layer = int(prev_layer.upper().replace('LAYER', '').strip())
            if isinstance(next_layer, str):
                next_layer = int(next_layer.upper().replace('LAYER', '').strip())

            if prev_layer == mid and next_layer == mid + 1:
                return entry['row']
        except (ValueError, AttributeError):
            continue

    return None


def swap_all_air_adhesive_pairs(layer_data, center_row):
    """
    Swap all air materials with their preceding Adhesive materials
    if air's row is smaller than center_row.

    Args:
        layer_data (list): Layer data list
        center_row (int): Center row number

    Returns:
        list: Modified layer data
    """
    swapped_count = 0

    for i in range(1, len(layer_data)):
        current_item = layer_data[i]
        previous_item = layer_data[i - 1]

        if (current_item.get('material') == 'air' and
                current_item.get('row', float('inf')) < center_row and
                'adhesive' in previous_item.get('material', '').lower()):
            layer_data[i], layer_data[i - 1] = layer_data[i - 1], layer_data[i]

            logger.info(f"Swapped air (row {current_item.get('row')}) with Adhesive")
            swapped_count += 1

    logger.info(f"Total swaps performed: {swapped_count}")
    return layer_data
